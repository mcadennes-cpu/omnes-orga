#!/usr/bin/env python3
"""
Compare le plan de roulement en base avec un fichier de roulement (.xlsx).

POURQUOI CET OUTIL
------------------
Ecrit le 01/08/2026 pendant l'etape 6. Il a servi a etablir trois constats qui
ont oriente toute la conception de MOD-1 :

  . le roulement en base suivait le fichier V1 (256 cases sur 283 identiques)
    et PAS le V2 (198 sur 305) -- le V2 n'avait jamais ete charge ;
  . la base avait derive de 27 cases, dont 14 n'etaient pas une derive mais un
    vestige de modelisation (les gardes de week-end de Dijon enregistrees sur
    « J3 Dijon » avant la creation du creneau « WE1 Dijon ») ;
  . le fichier V2 tranche les 13 divergences restantes, toutes en faveur du
    fichier -- l'optimiseur etant parti de l'Excel, pas de l'etat reel.

Il reste utile pour VERIFIER UN IMPORT (etape 6E) : apres avoir importe un
fichier, relancer cet outil doit donner zero ecart entre le plan et le fichier.

USAGE
-----
    python3 docs/sql/22-6-outil-comparer-roulement-fichiers.py                       # plan actif vs les 2 fichiers de docs/
    python3 docs/sql/22-6-outil-comparer-roulement-fichiers.py chemin/vers/plan.xlsx # un fichier precis
    python3 docs/sql/22-6-outil-comparer-roulement-fichiers.py --detail              # liste chaque ecart

Lecture seule : n'ecrit jamais en base.

DEPENDANCES : openpyxl (deja installe). Le jeton Supabase est lu dans le
trousseau macOS, comme pour tous les scripts du projet.
"""

import json
import re
import subprocess
import sys
import unicodedata
import urllib.request
from collections import Counter
from pathlib import Path

import openpyxl

RACINE = Path(__file__).resolve().parent.parent          # docs/
PROJET_SUPABASE = "ydihrgnixthrraprclox"                  # OMNES ORGA

# Le fichier ne contient que des initiales. La correspondance a ete deduite des
# regles de roulement en base (etape 7B-1), pas saisie a la main.
INITIALES = {
    "Mireille YUAN": "MY",
    "Thomas ETIENNE": "TE",
    "Xavier BAUDRILLART": "XB",
    "Airelle SAUVAGE": "AS",
    "Christophe BERTRAND": "CB",
    "Imane EL GARI": "IEG",
    "Caroline Chauvet": "CC",
    "Laurene DAUDIN": "LD",
    "Matthieu CADENNES": "MC",
}

JOURS = {1: "lundi", 2: "mardi", 3: "mercredi", 4: "jeudi",
         5: "vendredi", 6: "samedi", 0: "dimanche"}

# Creneau en base -> code du fichier. Le nom en base inclut le site ; les
# week-ends s'appellent « Garde » / « Doublon » dans le fichier.
# Source de verite : desiderata.yaml, section `correspondance_agenda`.
CRENEAU = {
    "J1 Beaune": "J1", "J2 Beaune": "J2", "J3 Beaune": "J3", "J4 Beaune": "J4",
    "J6 Beaune": "J6", "J7 Beaune": "J7", "J8 Beaune": "J8",
    "J2 Dijon": "J2", "J3 Dijon": "J3", "J4 Dijon": "J4", "J5 Dijon": "J5",
    "J6 Dijon": "J6", "J7 Dijon": "J7", "J8 Dijon": "J8",
    "J5 bis Dijon": "J5bis",
    "WE1 beaune 08h-20h": "Garde", "WE1 Dijon": "Garde",
    "WE2 beaune 08h-20h": "Doublon", "WE2 Dijon": "Doublon",
}


def sans_accent(valeur) -> str:
    """Normalise pour comparer : sans accent, sans espaces de bord."""
    texte = unicodedata.normalize("NFD", str(valeur or ""))
    return "".join(c for c in texte if unicodedata.category(c) != "Mn").strip()


def jeton_supabase() -> str:
    brut = subprocess.run(
        ["security", "find-generic-password", "-s", "Supabase CLI", "-w"],
        capture_output=True, text=True, check=True).stdout.strip()
    prefixe = "go-keyring-base64:"
    if brut.startswith(prefixe):
        import base64
        return base64.b64decode(brut[len(prefixe):]).decode().strip()
    return brut


def interroger(sql: str):
    # ⚠ User-Agent OBLIGATOIRE : l'API Management de Supabase repond 403 sur
    # celui que Python envoie par defaut (« Python-urllib/3.x »). Le jeton est
    # pourtant valide -- l'erreur est trompeuse.
    requete = urllib.request.Request(
        f"https://api.supabase.com/v1/projects/{PROJET_SUPABASE}/database/query",
        data=json.dumps({"query": sql}).encode(),
        headers={"Authorization": f"Bearer {jeton_supabase()}",
                 "Content-Type": "application/json",
                 "User-Agent": "omnes-orga-script/1.0"},
        method="POST")
    with urllib.request.urlopen(requete) as reponse:
        return json.load(reponse)


def lire_plan_actif():
    """Le plan en vigueur, sous forme {(semaine, jour, site, creneau): {initiales}}."""
    lignes = interroger("""
        select p.name as plan, r.rotation_week as sem, r.weekday as jour,
               si.name as site, st.name as creneau,
               pr.prenom || ' ' || pr.nom as medecin
          from agenda.rotation_plan_rules r
          join agenda.rotation_plans p on p.id = r.plan_id
          join agenda.sites si         on si.id = r.site_id
          join agenda.shift_types st   on st.id = r.shift_type_id
          join public.profiles pr      on pr.id = r.doctor_id
         where p.status = 'active'
    """)
    if isinstance(lignes, dict) and "message" in lignes:
        sys.exit(f"Erreur SQL : {lignes['message']}")

    grille, nom_plan, inconnus = {}, None, set()
    for ligne in lignes:
        nom_plan = ligne["plan"]
        ini = INITIALES.get(sans_accent(ligne["medecin"]))
        code = CRENEAU.get(sans_accent(ligne["creneau"]))
        if ini is None or code is None:
            inconnus.add((ligne["medecin"], ligne["creneau"]))
            continue
        cle = (ligne["sem"], JOURS[ligne["jour"]], sans_accent(ligne["site"]).lower(), code)
        grille.setdefault(cle, set()).add(ini)

    for medecin, creneau in sorted(inconnus):
        print(f"  ⚠ non reconnu : medecin={medecin!r} creneau={creneau!r}")
    return grille, nom_plan


def lire_fichier(chemin: Path):
    """
    Lit la grille d'un fichier de roulement.

    Repere la feuille et la ligne d'en-tete PAR LEUR CONTENU (une ligne
    contenant des libelles « S<n> <site> »), jamais par des coordonnees fixes :
    les deux fichiers du cabinet ont des dispositions differentes, et l'export
    Numbers duplique la feuille en la decalant d'une colonne.
    """
    classeur = openpyxl.load_workbook(chemin, data_only=True)

    for nom_feuille in classeur.sheetnames:
        feuille = classeur[nom_feuille]
        entete, colonnes = None, {}
        for ligne in range(1, min(feuille.max_row, 15) + 1):
            trouvees = {}
            for col in range(1, feuille.max_column + 1):
                valeur = sans_accent(feuille.cell(ligne, col).value)
                m = re.fullmatch(r"S(\d+)\s+(\w+)", valeur, re.I)
                if m:
                    trouvees[col] = (int(m.group(1)), m.group(2).lower())
            if len(trouvees) >= 4:
                entete, colonnes = ligne, trouvees
                break
        if not entete:
            continue

        # Les colonnes « jour » et « creneau » precedent la premiere semaine.
        premiere = min(colonnes)
        col_creneau, col_jour = premiere - 1, premiere - 2

        grille, jour = {}, None
        for ligne in range(entete + 1, feuille.max_row + 1):
            valeur_jour = sans_accent(feuille.cell(ligne, col_jour).value).lower()
            if valeur_jour:
                jour = valeur_jour          # le jour n'est ecrit que sur la 1re ligne du bloc
            creneau = sans_accent(feuille.cell(ligne, col_creneau).value)
            if not creneau or not jour:
                continue

            for col, (semaine, site) in colonnes.items():
                cellule = sans_accent(feuille.cell(ligne, col).value)
                if not cellule:
                    continue
                # Cellule composite du format V1 : « LD J7 » = medecin + creneau reel.
                morceaux = cellule.split()
                initiale = morceaux[0]
                code = morceaux[1] if len(morceaux) > 1 else creneau.split()[0]
                grille.setdefault((semaine, jour, site, code), set()).add(initiale)

        return grille, nom_feuille

    sys.exit(f"Aucune grille trouvee dans {chemin.name} (pas de libelle « S<n> <site> »).")


def comparer(gauche, droite, nom_gauche, nom_droite, detail=False):
    cles = set(gauche) | set(droite)
    identiques = [k for k in cles if gauche.get(k, set()) == droite.get(k, set())]
    gauche_seul = [k for k in cles if gauche.get(k, set()) and not droite.get(k, set())]
    droite_seul = [k for k in cles if droite.get(k, set()) and not gauche.get(k, set())]
    differents = [k for k in cles
                  if gauche.get(k, set()) and droite.get(k, set())
                  and gauche[k] != droite[k]]

    print(f"\n{nom_gauche}  vs  {nom_droite}")
    print(f"  affectations          : {sum(map(len, gauche.values()))} / {sum(map(len, droite.values()))}")
    print(f"  cases identiques      : {len(identiques)} / {len(cles)}")
    print(f"  seulement a gauche    : {len(gauche_seul)}")
    print(f"  seulement a droite    : {len(droite_seul)}")
    print(f"  medecin different     : {len(differents)}")

    if detail:
        for titre, cases in (("SEULEMENT A GAUCHE", gauche_seul),
                             ("SEULEMENT A DROITE", droite_seul),
                             ("MEDECIN DIFFERENT", differents)):
            if not cases:
                continue
            print(f"\n  --- {titre} ---")
            for sem, jour, site, code in sorted(cases):
                g = ",".join(sorted(gauche.get((sem, jour, site, code), set()))) or "-"
                d = ",".join(sorted(droite.get((sem, jour, site, code), set()))) or "-"
                print(f"    S{sem} {jour:<9} {site:<7} {code:<8} {g:<12} | {d}")

    return len(gauche_seul) + len(droite_seul) + len(differents)


def main():
    args = [a for a in sys.argv[1:] if a != "--detail"]
    detail = "--detail" in sys.argv

    fichiers = [Path(a) for a in args] or [
        RACINE / "planning-actuel_2025-12.xlsx",
        RACINE / "planning-V2_2026-07.xlsx",
    ]

    plan, nom_plan = lire_plan_actif()
    print(f"Plan actif en base : {nom_plan} — {sum(map(len, plan.values()))} affectations")

    total_ecarts = 0
    for chemin in fichiers:
        if not chemin.exists():
            print(f"\n⚠ introuvable : {chemin}")
            continue
        grille, feuille = lire_fichier(chemin)
        ecarts = comparer(plan, grille, "PLAN EN BASE", f"{chemin.name} [{feuille}]", detail)
        total_ecarts += ecarts

        repartition = Counter()
        for occupants in grille.values():
            for ini in occupants:
                repartition[ini] += 1
        print("  repartition du fichier :",
              ", ".join(f"{i}={n}" for i, n in sorted(repartition.items())))

    print(f"\nTotal des ecarts : {total_ecarts}")
    print("(apres un import reussi, l'ecart avec le fichier importe doit etre nul)")


if __name__ == "__main__":
    main()
