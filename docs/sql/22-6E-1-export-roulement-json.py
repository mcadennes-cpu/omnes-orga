#!/usr/bin/env python3
"""
Lit un fichier de roulement (.xlsx) et emet un JSON canonique pour l'import.

PLACE DANS LA CHAINE (cf. « MOD-1 bis » dans docs/integration-agenda.md)
-----------------------------------------------------------------------
    desiderata.yaml -> 1_optimize.py -> 2_generate_xlsx.py -> planning-Vx.xlsx
                                             (retouches manuelles Numbers)
                                                    verifie-planning.py
                                                    CE SCRIPT  ->  roulement-Vx.json
                                                    Omnes-Orga -> import -> rotation_plans

C'est le `3_export_app.py` de ce schema. Il vit ici pour l'instant, avec les
autres scripts du projet ; sa place definitive est le depot du pipeline Python,
aux cotes de `verifie-planning.py`.

POURQUOI PAR UN JSON, ET PAS UN PARSEUR .xlsx DANS L'APPLICATION
----------------------------------------------------------------
Le cabinet a produit deux fichiers de structures differentes en sept mois
(colonnes decalees, creneaux eclates, cellules composites, feuille dupliquee
par l'export Numbers). Toute cette fragilite de lecture reste donc du cote
Python, la ou vivent l'expertise et `verifie-planning.py`. L'application ne
recoit qu'un JSON deja normalise : son import devient une simple validation.

CE QUE LE JSON CONTIENT -- ET CE QU'IL NE CONTIENT PAS
-------------------------------------------------------
Il contient l'image FIDELE du fichier : les codes tels qu'ils y sont ecrits
(« CB », « J2 », « Beaune »), rien de plus.

Il ne contient AUCUN identifiant de la base, aucun nom complet de medecin,
aucun horaire de creneau. C'est volontaire : la resolution code -> compte
medecin / site / shift_type est le travail de l'ecran de correspondance de
l'application (6E-3), qui la memorise d'un import a l'autre. Recopier ici les
tables de `desiderata.yaml` en ferait une troisieme source de verite, vouee a
diverger -- c'est le defaut deja releve sur `1_optimize.py`.

Frontiere : PYTHON LIT LE FICHIER, L'APPLICATION RESOUT LES IDENTITES.

USAGE
-----
    python3 docs/sql/22-6E-1-export-roulement-json.py docs/planning-V2_2026-07.xlsx \\
        --nom "Roulement V2 - 9 associes" --date-debut 2027-01-04

    -o chemin.json   sortie explicite (defaut : « roulement-*.json » a cote du .xlsx)
    --detail         liste toutes les affectations lues

La date de debut doit etre un LUNDI : la base refuse tout plan dont le
`start_date` n'en est pas un (une autre date decalerait silencieusement toute
la numerotation des semaines). Le controle est fait ici pour ne pas decouvrir
l'erreur au moment de l'import.

Lecture seule : n'ecrit jamais en base, ne touche jamais au .xlsx.

DEPENDANCES : openpyxl (deja installe). Aucune autre.
"""

import argparse
import datetime as dt
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

import openpyxl

# Les sept jours, dans l'ordre de la semaine. La cle est la forme normalisee
# (sans accent, minuscules) ; la valeur est la forme canonique du JSON.
JOURS = {
    "lundi": "Lundi", "mardi": "Mardi", "mercredi": "Mercredi",
    "jeudi": "Jeudi", "vendredi": "Vendredi", "samedi": "Samedi",
    "dimanche": "Dimanche",
}
ORDRE_JOURS = {nom: i for i, nom in enumerate(JOURS.values())}

# Un code medecin est un sigle court en majuscules : « MC », « IEG ».
# Sert uniquement a signaler une cellule qui n'y ressemble pas (note de bas de
# page, commentaire) -- jamais a rejeter une valeur.
FORME_CODE_MEDECIN = re.compile(r"^[A-Z]{1,4}$")


def sans_accent(valeur) -> str:
    """Normalise pour comparer : sans accent, sans espaces de bord."""
    texte = unicodedata.normalize("NFD", str(valeur or ""))
    return "".join(c for c in texte if unicodedata.category(c) != "Mn").strip()


class Anomalies:
    """Rapport d'anomalies -- collecte, puis affichage groupe par type."""

    def __init__(self):
        self.liste = []

    def ajouter(self, type_, message, ou=None):
        self.liste.append({"type": type_, "message": message, "ou": ou})

    def afficher(self):
        if not self.liste:
            print("\nAnomalies : aucune.")
            return
        groupes = defaultdict(list)
        for a in self.liste:
            groupes[a["type"]].append(a)
        print(f"\nAnomalies : {len(self.liste)}")
        for type_, items in sorted(groupes.items()):
            print(f"  [{type_}] {len(items)}")
            for a in items[:10]:
                ou = f" ({a['ou']})" if a["ou"] else ""
                print(f"      {a['message']}{ou}")
            if len(items) > 10:
                print(f"      ... et {len(items) - 10} autres")


def reperer_grille(feuille):
    """
    Cherche la ligne d'en-tete PAR SON CONTENU : celle qui porte au moins
    quatre libelles « S<n> <site> ». Jamais par des coordonnees fixes -- les
    deux fichiers du cabinet ont des dispositions differentes et l'export
    Numbers duplique la feuille en la decalant d'une colonne.

    Retourne (ligne_entete, {colonne: (semaine, site)}) ou (None, {}).
    """
    for ligne in range(1, min(feuille.max_row, 15) + 1):
        trouvees = {}
        for col in range(1, feuille.max_column + 1):
            valeur = sans_accent(feuille.cell(ligne, col).value)
            m = re.fullmatch(r"S(\d+)\s+(\w+)", valeur, re.I)
            if m:
                site = m.group(2)
                trouvees[col] = (int(m.group(1)), site[0].upper() + site[1:].lower())
        if len(trouvees) >= 4:
            return ligne, trouvees
    return None, {}


def lire_fichier(chemin: Path, anomalies: Anomalies):
    """
    Lit la grille de roulement et retourne (affectations, nom_feuille).

    Gere les deux formats du cabinet : creneaux eclates une ligne chacun (V2)
    et lignes composites « J6 ou J7 ou J8 » dont la cellule porte le creneau
    reellement retenu (V1, « LD J7 »).
    """
    classeur = openpyxl.load_workbook(chemin, data_only=True)

    retenue, autres = None, []
    for nom_feuille in classeur.sheetnames:
        entete, colonnes = reperer_grille(classeur[nom_feuille])
        if not entete:
            continue
        if retenue is None:
            retenue = (nom_feuille, entete, colonnes)
        else:
            autres.append(nom_feuille)

    if retenue is None:
        sys.exit(f"Aucune grille trouvee dans {chemin.name} "
                 f"(aucune ligne ne porte de libelle « S<n> <site> »).")

    nom_feuille, entete, colonnes = retenue
    for ignoree in autres:
        # C'est le doublon « Feuille 1-1 » de l'export Numbers. On le signale
        # plutot que de le lire en silence : si un jour les deux divergeaient,
        # il faudrait le savoir.
        anomalies.ajouter("feuille_ignoree",
                          f"la feuille {ignoree!r} porte aussi une grille, elle n'a pas ete lue")

    feuille = classeur[nom_feuille]
    premiere = min(colonnes)
    col_creneau, col_jour = premiere - 1, premiere - 2
    if col_jour < 1:
        sys.exit(f"Grille de {nom_feuille!r} : pas de place pour les colonnes "
                 f"jour et creneau avant la colonne {premiere}.")

    affectations, occupants = [], defaultdict(set)
    jour = None

    for ligne in range(entete + 1, feuille.max_row + 1):
        valeur_jour = sans_accent(feuille.cell(ligne, col_jour).value).lower()
        if valeur_jour:
            # Le jour n'est ecrit que sur la premiere ligne du bloc : on le
            # propage vers le bas jusqu'au prochain.
            if valeur_jour not in JOURS:
                anomalies.ajouter("jour_inconnu",
                                  f"libelle de jour non reconnu : {valeur_jour!r}",
                                  f"ligne {ligne}")
                jour = None
                continue
            jour = JOURS[valeur_jour]

        libelle_creneau = sans_accent(feuille.cell(ligne, col_creneau).value)
        cellules = {c: sans_accent(feuille.cell(ligne, c).value) for c in colonnes}
        remplies = {c: v for c, v in cellules.items() if v}

        if not libelle_creneau or not jour:
            if remplies:
                manque = "aucun creneau" if not libelle_creneau else "aucun jour"
                anomalies.ajouter("ligne_ignoree",
                                  f"{len(remplies)} cellule(s) remplie(s) mais {manque} "
                                  f"sur cette ligne",
                                  f"ligne {ligne}")
            continue

        # Libelle a plusieurs mots (« J6 ou J7 ou J8 ») : le creneau ne se
        # deduit pas de la ligne, chaque cellule doit porter le sien.
        libelle_ambigu = len(libelle_creneau.split()) > 1

        for col, cellule in remplies.items():
            semaine, site = colonnes[col]
            ou = f"S{semaine} {jour} {site} (ligne {ligne})"
            morceaux = cellule.split()

            if len(morceaux) > 2:
                anomalies.ajouter("cellule_ambigue",
                                  f"cellule a plus de deux valeurs : {cellule!r} -- ignoree", ou)
                continue

            medecin = morceaux[0]
            if len(morceaux) == 2:
                creneau = morceaux[1]
            elif libelle_ambigu:
                anomalies.ajouter("creneau_ambigu",
                                  f"la ligne {libelle_creneau!r} ne designe pas un creneau unique "
                                  f"et la cellule {cellule!r} n'en precise aucun -- ignoree", ou)
                continue
            else:
                creneau = libelle_creneau

            if not FORME_CODE_MEDECIN.match(medecin):
                anomalies.ajouter("code_medecin_suspect",
                                  f"{medecin!r} ne ressemble pas a des initiales "
                                  f"-- conserve, a verifier a l'ecran de correspondance", ou)

            cle = (semaine, jour, site, creneau)
            if medecin in occupants[cle]:
                # La base refuse le meme medecin deux fois sur une case (mais
                # accepte deux medecins differents : c'est le « Doublon »).
                anomalies.ajouter("doublon_exact",
                                  f"{medecin} apparait deux fois sur la meme case "
                                  f"{creneau} -- conserve une seule fois", ou)
                continue
            occupants[cle].add(medecin)

            affectations.append({"medecin": medecin, "semaine": semaine,
                                 "jour": jour, "site": site, "creneau": creneau})

    return affectations, nom_feuille


def chemin_sortie(source: Path, demande: Optional[str]) -> Path:
    if demande:
        return Path(demande)
    nom = source.stem
    nom = nom[len("planning-"):] if nom.startswith("planning-") else nom
    return source.with_name(f"roulement-{nom}.json")


def main():
    analyseur = argparse.ArgumentParser(
        description="Convertit un fichier de roulement .xlsx en JSON canonique.")
    analyseur.add_argument("fichier", help="le fichier de roulement (.xlsx)")
    analyseur.add_argument("--nom", required=True,
                           help="nom du plan, ex. \"Roulement V2 - 9 associes\"")
    analyseur.add_argument("--date-debut", required=True,
                           help="premier lundi du cycle, au format AAAA-MM-JJ")
    analyseur.add_argument("-o", "--sortie", help="chemin du JSON produit")
    analyseur.add_argument("--detail", action="store_true",
                           help="liste toutes les affectations lues")
    args = analyseur.parse_args()

    source = Path(args.fichier)
    if not source.exists():
        sys.exit(f"Fichier introuvable : {source}")

    try:
        date_debut = dt.date.fromisoformat(args.date_debut)
    except ValueError:
        sys.exit(f"Date invalide : {args.date_debut!r} -- format attendu AAAA-MM-JJ.")
    if date_debut.weekday() != 0:
        jour_fr = list(JOURS.values())[date_debut.weekday()]
        sys.exit(f"--date-debut doit etre un LUNDI. Le {date_debut} est un {jour_fr}.\n"
                 f"    Lundi le plus proche : {date_debut - dt.timedelta(days=date_debut.weekday())}\n"
                 f"    (la base refuse tout plan dont le start_date n'est pas un lundi : "
                 f"toute la numerotation des semaines en depend.)")

    anomalies = Anomalies()
    affectations, nom_feuille = lire_fichier(source, anomalies)
    if not affectations:
        sys.exit(f"Aucune affectation lue dans {source.name}.")

    semaines = sorted({a["semaine"] for a in affectations})
    cycle = max(semaines)
    if semaines != list(range(1, cycle + 1)):
        manquantes = [s for s in range(1, cycle + 1) if s not in semaines]
        anomalies.ajouter("semaine_vide",
                          f"aucune affectation en semaine(s) {manquantes} "
                          f"alors que le cycle en compte {cycle}")

    affectations.sort(key=lambda a: (a["semaine"], ORDRE_JOURS[a["jour"]],
                                     a["site"], a["creneau"], a["medecin"]))

    document = {
        "plan": {
            "nom": args.nom,
            "cycle_semaines": cycle,
            "date_debut": date_debut.isoformat(),
            "source": source.name,
            "feuille": nom_feuille,
            "exporte_le": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        },
        "sites": sorted({a["site"] for a in affectations}),
        "creneaux": sorted({a["creneau"] for a in affectations}),
        "medecins": sorted({a["medecin"] for a in affectations}),
        "jours": sorted({a["jour"] for a in affectations}, key=ORDRE_JOURS.get),
        "affectations": affectations,
        "anomalies": anomalies.liste,
    }

    sortie = chemin_sortie(source, args.sortie)
    sortie.write_text(json.dumps(document, ensure_ascii=False, indent=2) + "\n",
                      encoding="utf-8")

    print(f"{source.name}  [feuille {nom_feuille!r}]")
    print(f"  cycle             : {cycle} semaines, debut le {date_debut} (lundi)")
    print(f"  sites             : {', '.join(document['sites'])}")
    print(f"  creneaux          : {', '.join(document['creneaux'])}")
    print(f"  affectations      : {len(affectations)}")

    repartition = Counter(a["medecin"] for a in affectations)
    print(f"  medecins          : {len(repartition)}")
    print("  repartition       :",
          ", ".join(f"{m}={n}" for m, n in sorted(repartition.items())))

    if args.detail:
        print("\n  --- AFFECTATIONS ---")
        for a in affectations:
            print(f"    S{a['semaine']} {a['jour']:<9} {a['site']:<7} "
                  f"{a['creneau']:<8} {a['medecin']}")

    anomalies.afficher()
    print(f"\nEcrit : {sortie}")
    print("Controle : apres import, "
          "22-6-outil-comparer-roulement-fichiers.py doit donner zero ecart.")


if __name__ == "__main__":
    main()
